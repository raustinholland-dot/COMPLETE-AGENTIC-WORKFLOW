#!/usr/bin/env python3
"""
Ingest Clearwater internal scoping documents into the clearwater_internal Qdrant collection.
Extracts text from xlsx/docx files, chunks, embeds via OpenAI, upserts to Qdrant.

Usage: python3 scripts/ingest-clearwater-internal.py
"""

import os
import sys
import json
import uuid
import time
import requests
import openpyxl
import docx

# ── Config ────────────────────────────────────────────────────────────────────
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "REDACTED")
QDRANT_URL = "http://localhost:6333"
COLLECTION  = "clearwater_internal"
CHUNK_SIZE  = 1800
OVERLAP     = 180
EMBED_MODEL = "text-embedding-3-small"

# ── Source files ──────────────────────────────────────────────────────────────
BASE = "/Users/austinhollsnd/Desktop/Sales Team Support Templates copy/Scoping Questions"

FILES = [
    {
        "path": f"{BASE}/Clearwater Scoping Questions_SECURITY ENGINEERING SERVICES_Jeremy Hughes_20250624.xlsx",
        "service_line": "security_engineering",
        "doc_type": "scoping_template",
        "label": "Security Engineering Services",
    },
    {
        "path": f"{BASE}/Clearwater Scoping Questions_PRIVACY-COMPLIANCE-AUDIT_20240927.xlsx",
        "service_line": "privacy_compliance_audit",
        "doc_type": "scoping_template",
        "label": "Privacy, Compliance & Audit",
    },
    {
        "path": f"{BASE}/Clearwater Scoping Questions_CLEARWATER MANAGED SECURITY-HEALTHCARE_AKERS-013125.xlsx",
        "service_line": "managed_security_healthcare",
        "doc_type": "scoping_template",
        "label": "Managed Security Services (Healthcare)",
    },
    {
        "path": f"{BASE}/Clearwater Scoping Questions_CYBER RESILIENCE STRATEGIC ADVISORS_BIA & BCP_Heather Hanson_20250610.docx",
        "service_line": "bia_bcp",
        "doc_type": "scoping_template",
        "label": "BIA & BCP (Cyber Resilience)",
    },
    {
        "path": None,  # inline text — no file
        "service_line": "managed_azure_cloud",
        "doc_type": "discovery_call_template",
        "label": "Azure Cloud for Healthcare — Discovery Call Template",
        "source_file": "Azure Cloud for Healthcare Discovery Call Template (Steve Akers)",
        "inline_text": """[INTRODUCTIONS & EXPECTATIONS]
Please introduce your role and involvement in cloud or IT transformation.
What would make today's conversation a success for you?

[STRATEGIC GOALS & BUSINESS DRIVERS]
What are your key strategic initiatives?
Is patient outcome improvement a current priority?
Is telehealth expansion a current priority?
Is operational efficiency a current priority?
Is regulatory readiness a current priority?
Is cost control a current priority?
How does your organization view cloud adoption — as a risk, opportunity, or necessity?
Is there an executive mandate for digital transformation or a cloud-first strategy?

[CURRENT ENVIRONMENT & CLINICAL SYSTEMS]
Are you currently using Azure or any other cloud platforms?
What is your current EHR/EMR platform (e.g., Epic, Cerner, Meditech)?
Which clinical, administrative, or imaging workloads are on-prem vs. cloud?
Are you integrating with any Health Information Exchanges (HIEs)?
How are patient records accessed across locations or departments?

[INFRASTRUCTURE, SECURITY & COMPLIANCE]
How do you currently meet HIPAA and HITRUST compliance in your IT environment?
Are you leveraging Azure tools like Microsoft Defender for Cloud, Azure Policy, or Sentinel?
Do you have multi-tenant concerns (e.g., shared environments across facilities)?
How is identity and access managed (e.g., Azure AD, hybrid AD, SSO)?
How do you protect ePHI across storage, apps, and endpoints?
Is there a Business Associate Agreement (BAA) in place with Microsoft?

[WORKLOADS & MODERNIZATION]
Are there any legacy systems (RIS, PACS, lab systems) targeted for cloud migration?
Would app modernization (Azure App Services or containers) help with scalability or compliance?
Is interoperability a key objective — for example, integrating third-party systems or partners?
Are AI/ML for diagnostics a use case of interest?
Are clinical analytics or population health dashboards a use case of interest?
Is IoMT integration a use case of interest?

[PROCUREMENT, BUDGET & SUPPORT]
Are you procuring Azure via EA, CSP, or other channels?
Are cloud costs allocated by facility, department, or centralized IT?
Do you require a 24/7 support model due to patient-critical services?

[TIMELINE, STAKEHOLDERS & NEXT STEPS]
What is your target timeline for Azure migration or optimization?
Who are the clinical, technical, and compliance stakeholders?
What would success look like for this initiative?

[OPTIONAL ADD-ONS TO DISCUSS]
Epic on Azure Strategy
Clinical Downtime Resilience Plans
Azure for Imaging & Storage (DICOM, VNA, etc.)
FHIR & HL7 Integration Readiness""",
    },
]


# ── Text extraction ───────────────────────────────────────────────────────────

def extract_xlsx(path: str) -> str:
    """Extract all cell text from all sheets in an xlsx file."""
    wb = openpyxl.load_workbook(path, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_text = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip() not in ("", "None")]
            if cells:
                sheet_text.append(" | ".join(cells))
        if sheet_text:
            parts.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(sheet_text))
    return "\n\n".join(parts)


def extract_docx(path: str) -> str:
    """Extract all paragraph text from a docx file."""
    doc = docx.Document(path)
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def extract_text(file_info: dict) -> str:
    path = file_info.get("path")
    if file_info.get("inline_text"):
        text = file_info["inline_text"]
        source_file = file_info.get("source_file", "inline")
    elif path and path.endswith(".xlsx"):
        text = extract_xlsx(path)
        source_file = os.path.basename(path)
    elif path and path.endswith(".docx"):
        text = extract_docx(path)
        source_file = os.path.basename(path)
    else:
        raise ValueError(f"Unsupported file type: {path}")

    # Prepend context header (mirrors CW-01 contextual enrichment pattern)
    header = (
        f"[CLEARWATER INTERNAL DOCUMENT]\n"
        f"Document Type: {file_info['doc_type']}\n"
        f"Service Line: {file_info['service_line']}\n"
        f"Label: {file_info['label']}\n"
        f"Source File: {source_file}\n"
        f"[END HEADER]\n\n"
    )
    return header + text


# ── Chunking ──────────────────────────────────────────────────────────────────

def chunk_text(text: str, chunk_size: int = CHUNK_SIZE, overlap: int = OVERLAP) -> list[str]:
    """Fixed-size character chunking with overlap."""
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]
        if chunk.strip():
            chunks.append(chunk)
        start += chunk_size - overlap
    return chunks


# ── Embedding ─────────────────────────────────────────────────────────────────

def embed(text: str) -> list[float]:
    resp = requests.post(
        "https://api.openai.com/v1/embeddings",
        headers={"Authorization": f"Bearer {OPENAI_API_KEY}", "Content-Type": "application/json"},
        json={"model": EMBED_MODEL, "input": text},
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()["data"][0]["embedding"]


# ── Qdrant upsert ─────────────────────────────────────────────────────────────

def upsert_points(points: list[dict]):
    resp = requests.put(
        f"{QDRANT_URL}/collections/{COLLECTION}/points",
        headers={"Content-Type": "application/json"},
        json={"points": points},
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    # Verify collection exists
    r = requests.get(f"{QDRANT_URL}/collections/{COLLECTION}", timeout=10)
    if r.status_code != 200:
        print(f"ERROR: Collection '{COLLECTION}' not found in Qdrant. Run create-qdrant-collection.sh first.")
        sys.exit(1)
    print(f"Collection '{COLLECTION}' found. Starting ingestion...\n")

    total_chunks = 0

    for file_info in FILES:
        path = file_info["path"]
        label = file_info["label"]

        is_inline = bool(file_info.get("inline_text"))
        if not is_inline and (not path or not os.path.exists(path)):
            print(f"  SKIP (file not found): {path}")
            continue

        print(f"Processing: {label}")
        print(f"  Source: {'inline text' if is_inline else os.path.basename(path)}")

        # Extract
        text = extract_text(file_info)
        print(f"  Extracted: {len(text):,} chars")

        # Chunk
        chunks = chunk_text(text)
        print(f"  Chunks: {len(chunks)}")

        # Embed + upsert in batches of 10
        points = []
        for i, chunk in enumerate(chunks):
            vector = embed(chunk)
            points.append({
                "id": str(uuid.uuid4()),
                "vector": vector,
                "payload": {
                    "service_line": file_info["service_line"],
                    "doc_type": file_info["doc_type"],
                    "label": label,
                    "source_file": file_info.get("source_file", os.path.basename(path) if path else "inline"),
                    "chunk_index": i,
                    "page_content": chunk,
                },
            })

            # Upsert in batches of 10
            if len(points) >= 10:
                upsert_points(points)
                print(f"  Upserted chunks {i - 8}-{i + 1}...")
                points = []
                time.sleep(0.2)  # brief pause to avoid rate limits

        # Flush remaining
        if points:
            upsert_points(points)
            print(f"  Upserted final {len(points)} chunk(s)")

        total_chunks += len(chunks)
        print(f"  Done.\n")

    print(f"Ingestion complete. Total chunks upserted: {total_chunks}")

    # Verify
    r = requests.get(f"{QDRANT_URL}/collections/{COLLECTION}", timeout=10)
    count = r.json()["result"]["points_count"]
    print(f"Qdrant '{COLLECTION}' collection now has {count} vectors.")


if __name__ == "__main__":
    main()
